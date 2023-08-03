from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

FUNCTIONAL_DISABILITIES_NAMES = ["seeing", "hearing", "walking", "cognition", "self-care", "communication"]
DISABILITIES_QUESTION_TYPES = {
    "Functional disability": "1",
    "Non functional disability": "2",
    "Broad activity limitation": "3",
    "Activities of Daily Living": "4",
    "Disability in general": "5",
    "Other type of disabilities": "6",
}


def format_data(data):
    sheet = list()
    for filename in data:
        column = [Path(filename).name,
                  data[filename]["dataset_country"],
                  data[filename]["dataset_name"],
                  data[filename]["dataset_type"] + " - " + data[filename]["dataset_years"],
                  data[filename]["dataset_reviewer"],
                  ""
                  ]
        functional_disabilities_values = get_functional_disabilities(data[filename]["disabilities"])
        column.extend(functional_disabilities_values)
        scales_values = get_scales_values(data[filename]["disabilities"])
        column.extend(scales_values)
        column.extend(["", "", "", ""])
        column.extend(compute_sums(column))
        column.append(compute_pages(data[filename]["disabilities"]))
        column.append("")
        column.extend(get_other_types_disability(data[filename]["disabilities"],
                                                 DISABILITIES_QUESTION_TYPES["Functional disability"]))
        column.append("")
        column.extend(get_other_types_disability(data[filename]["disabilities"],
                                                 DISABILITIES_QUESTION_TYPES["Broad activity limitation"]))
        column.extend(get_other_types_disability(data[filename]["disabilities"],
                                                 DISABILITIES_QUESTION_TYPES["Activities of Daily Living"]))
        column.extend(get_other_types_disability(data[filename]["disabilities"],
                                                 DISABILITIES_QUESTION_TYPES["Disability in general"]))
        other = get_other_types_disability(data[filename]["disabilities"],
                                           DISABILITIES_QUESTION_TYPES["Other type of disabilities"])
        non_functional = get_other_types_disability(data[filename]["disabilities"],
                                                    DISABILITIES_QUESTION_TYPES["Non functional disability"])

        texts = [other[1], non_functional[1]]
        texts = [i for i in texts if i]
        column.extend([str(max(int(other[0]), int(non_functional[0]))), ";".join(texts)])
        column.append("")
        column.append(str(int(column[26]) + int(column[27]) + int(column[35]) + int(column[37]) + int(column[39]) + int(
            column[41])))
        any_disability = "1" if data[filename]["disabilities"] else "0"
        column.append(any_disability)
        column.append(compile_disability_types(column))
        column.append(data[filename]["dataset_comments"])

        sheet.append(column)
    return sheet


def compile_disability_types(values):
    type_list = list()
    if values[26] == "1":
        type_list.append("(1)")
    if values[27] == "1":
        type_list.append("(2)")
    if values[35] == "1":
        type_list.append("(3)")
    if values[37] == "1":
        type_list.append("(4)")
    if values[39] == "1":
        type_list.append("(5)")
    if values[41] == "1":
        type_list.append("(6)")
    return "".join(type_list)


def get_other_types_disability(data, type):
    disability = "0"
    texts = list()
    if "other" in data:
        for item in data["other"]:
            if item["question_type"] == type:
                disability = "1"
                texts.append(item["texts"])
        return [disability, ";".join(texts)]
    return [disability, ""]


def compute_sums(values):
    wgss_sum = compute_wg_ss_sum(values)
    disability_sum = compute_disability_sum(values)
    is_wgss_present = "1" if wgss_sum >= 5 else "0"
    is_other_disability_present = "1" if disability_sum >= 4 else "0"
    return [str(wgss_sum), str(disability_sum), is_wgss_present, is_other_disability_present, "", ""]


def compute_pages(data):
    page_list = list()
    for disability in FUNCTIONAL_DISABILITIES_NAMES:
        if disability in data:
            for item in data[disability]:
                page_list.append(int(item["page_number"]))
    page_list = sorted(page_list)
    page_list = [str(page) for page in page_list]
    return "pages : " + ";".join(page_list)


def get_functional_disabilities(data):
    values = list()
    for disability in FUNCTIONAL_DISABILITIES_NAMES:
        if disability in data:
            values.append("1")
            wording = "0"
            for item in data[disability]:
                if item["wg_wording"] == "1":
                    wording = "1"
            if wording == "0":
                texts = ";".join([item["texts"] for item in data[disability]])
                values.append(texts)
            else:
                values.append("1")
        else:
            values.append("0")
            values.append("")
    return values


def get_scales_values(data):
    values = list()
    scales = "0"
    scales_text = ""
    for disability in data:
        for item in data[disability]:
            if item["wg_scale"] == "1":
                scales = "1"
            elif item["wg_scale_text"]:
                scales_text = item["wg_scale_text"]
    if scales == "1":
        values.append("1")
        values.append("")
    else:
        values.append("0")
        values.append(scales_text)
    return values


def compute_wg_ss_sum(values):
    total = 0
    for i in range(6):
        if values[i * 2 + 7] == "1":
            total += 1
    if values[18] == "1":
        total += 1
    return total


def compute_disability_sum(values):
    total = 0
    for i in range(4):
        if values[i * 2 + 6] == "1":
            total += 1
    return total


def write_excel_report(matrix):
    wb = load_workbook("static/media/Empty_log2.xlsx")
    sheet = wb["Log"]
    for i in range(len(matrix)):
        for j in range(len(matrix[i])):
            sheet.cell(column=i + 2, row=j + 1).value = matrix[i][j]
    mem = BytesIO()
    wb.save(mem)
    mem.seek(0)
    return mem


def generate_log(data):
    matrix = format_data(data)
    file = write_excel_report(matrix)
    return file
